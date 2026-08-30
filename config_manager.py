"""
config_manager.py - Gerenciamento da planilha Excel (agenda_esf.xlsx)
Abas: Config, Agenda, Receitas, Log
Otimizado com cache em memória baseado em mtime e I/O thread-safe.
"""
import json
import os
import time
import threading
import pandas as pd
from openpyxl import Workbook
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime, timedelta
from utils import now, timestamp_str, date_str, normalize_text, get_modalidade_rules, search_match

# Constantes
ARQUIVO_EXCEL = "agenda_esf.xlsx"
COLUNAS_AGENDA = [
    'id', 'paciente', 'medico', 'tipo_consulta', 'data', 'hora',
    'encaixe', 'status', 'compareceu', 'quando_confirmou_presenca',
    'observacao', 'criado_em', 'criado_por'
]
COLUNAS_RECEITAS = [
    'id', 'paciente', 'status', 'quem_retirou', 'observacao',
    'data_pedido', 'data_retirada'
]
COLUNAS_LOG = [
    'id', 'timestamp', 'acao', 'detalhes', 'usuario'
]
COLUNAS_CONFIG = [
    'tipo', 'chave', 'valor'
]

# Tipos de ação para o Log
ACAO_AGENDAR = 'AGENDAR'
ACAO_CANCELAR = 'CANCELAR'
ACAO_ENCAIXAR = 'ENCAIXAR'
ACAO_EDITAR = 'EDITAR'
ACAO_PEDIR_RECEITA = 'PEDIR_RECEITA'
ACAO_RETIRAR_RECEITA = 'RETIRAR_RECEITA'
ACAO_MARCAR_PRESENCA = 'MARCAR_PRESENCA'
ACAO_EDITAR_CONFIG = 'EDITAR_CONFIG'
ACAO_DESMARCAR_PRESENCA = 'DESMARCAR_PRESENCA'

MODALIDADES_DISPONIVEIS = ['Normal', 'Domiciliar', 'GERCON', 'Criança', 'Gestante']

# Dados iniciais para a aba Config
_CONFIG_INICIAL = [
    ('modalidade', 'domiciliar', 'Domiciliar - Qui 08:00-12:00'),
    ('modalidade', 'gercon', 'GERCON - Qua 08:00-12:00'),
    ('modalidade', 'crianca', 'Criança - Qua 13:00-17:00'),
    ('modalidade', 'gestante', 'Gestante - Ter 08:00-12:00'),
]


class ExcelManager:
    """Gerencia todas as operações de leitura/escrita na planilha Excel com cache inteligente."""
    
    def __init__(self, caminho: str = None):
        self.caminho = caminho or ARQUIVO_EXCEL
        self._lock = threading.RLock()
        self._cache_mtime = 0.0
        self._cache_dfs: Dict[str, pd.DataFrame] = {}
        self._inicializar()
    
    def _inicializar(self):
        """Cria o arquivo Excel com as abas necessárias se não existir"""
        with self._lock:
            if not os.path.exists(self.caminho):
                self._criar_planilha()
    
    def _criar_planilha(self):
        """Cria planilha do zero com estrutura inicial"""
        with self._lock:
            wb = Workbook()
            
            ws_config = wb.active
            ws_config.title = "Config"
            ws_config.append(COLUNAS_CONFIG)
            for linha in _CONFIG_INICIAL:
                ws_config.append(linha)
            for col in ['A', 'B', 'C']:
                ws_config.column_dimensions[col].width = 20
            
            ws_agenda = wb.create_sheet("Agenda")
            ws_agenda.append(COLUNAS_AGENDA)
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                ws_agenda.column_dimensions[col].width = 18
            
            ws_receitas = wb.create_sheet("Receitas")
            ws_receitas.append(COLUNAS_RECEITAS)
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws_receitas.column_dimensions[col].width = 18
            
            ws_log = wb.create_sheet("Log")
            ws_log.append(COLUNAS_LOG)
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws_log.column_dimensions[col].width = 22
            
            pasta = os.path.dirname(self.caminho)
            if pasta and not os.path.exists(pasta):
                os.makedirs(pasta, exist_ok=True)
            wb.save(self.caminho)
            
            try:
                self._cache_mtime = os.path.getmtime(self.caminho)
            except OSError:
                self._cache_mtime = time.time()
            self._cache_dfs.clear()

    def _obter_sheet(self, sheet_name: str, colunas_default: Optional[List[str]] = None) -> pd.DataFrame:
        """
        Lê a aba da planilha utilizando cache em memória validado pelo mtime do arquivo.
        Se o arquivo não foi alterado em disco, retorna do cache em <0.1ms.
        """
        with self._lock:
            if not os.path.exists(self.caminho):
                self._criar_planilha()
            
            try:
                mtime_atual = os.path.getmtime(self.caminho)
            except OSError:
                mtime_atual = 0.0
            
            # Se o arquivo foi modificado externamente, limpa o cache
            if mtime_atual != self._cache_mtime:
                self._cache_dfs.clear()
                self._cache_mtime = mtime_atual
            
            if sheet_name in self._cache_dfs:
                return self._cache_dfs[sheet_name].copy()
            
            try:
                df = pd.read_excel(self.caminho, sheet_name=sheet_name, dtype=str)
            except Exception:
                df = pd.DataFrame(columns=colunas_default or [])
            
            if colunas_default:
                for col in colunas_default:
                    if col not in df.columns:
                        df[col] = ''
            
            self._cache_dfs[sheet_name] = df
            return df.copy()

    def _salvar_sheet(self, df: pd.DataFrame, sheet_name: str, max_retries: int = 5, delay: float = 0.4):
        """
        Salva o DataFrame na aba especificada, atualizando o cache em memória
        e gravando no arquivo com retentativas automáticas e proteção de concorrência.
        """
        with self._lock:
            # Atualiza cache em memória imediatamente
            self._cache_dfs[sheet_name] = df.copy()
            
            ultimo_erro = None
            for tentativa in range(1, max_retries + 1):
                try:
                    with pd.ExcelWriter(self.caminho, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                    try:
                        self._cache_mtime = os.path.getmtime(self.caminho)
                    except OSError:
                        pass
                    return
                except (PermissionError, OSError) as e:
                    ultimo_erro = e
                    if tentativa < max_retries:
                        time.sleep(delay)
            
            nome_arq = os.path.basename(self.caminho)
            raise PermissionError(
                f"Não foi possível salvar na planilha '{nome_arq}'.\n"
                f"O arquivo está aberto no Microsoft Excel ou sincronizando no OneDrive/Google Drive.\n"
                f"Por favor, feche a planilha no Excel e tente novamente."
            ) from ultimo_erro

    # ==================== CONFIG ====================

    def _garantir_config_basica(self) -> pd.DataFrame:
        """Garante que a aba Config exista e contenha os registros básicos necessários."""
        with self._lock:
            df = self._obter_sheet("Config", COLUNAS_CONFIG)

            if df.empty:
                df = pd.DataFrame(columns=COLUNAS_CONFIG)

            col_missing = False
            for col in COLUNAS_CONFIG:
                if col not in df.columns:
                    df[col] = ''
                    col_missing = True

            df = df.copy()
            df = df.dropna(how='all')
            df['tipo'] = df['tipo'].fillna('').astype(str)
            df['chave'] = df['chave'].fillna('').astype(str)
            df['valor'] = df['valor'].fillna('').astype(str)

            linhas_para_adicionar = []
            existente = set()
            for _, linha in df.iterrows():
                chave = (str(linha.get('tipo', '')).strip(), str(linha.get('chave', '')).strip())
                existente.add(chave)

            # Se estiver vazio, adiciona configuração padrão
            if df.empty:
                for tipo, chave, valor in _CONFIG_INICIAL:
                    key = (str(tipo).strip(), str(chave).strip())
                    if key not in existente:
                        linhas_para_adicionar.append({'tipo': tipo, 'chave': chave, 'valor': valor})
                        existente.add(key)

            if linhas_para_adicionar:
                df = pd.concat([df, pd.DataFrame(linhas_para_adicionar)], ignore_index=True)
                df = df.drop_duplicates(subset=['tipo', 'chave'], keep='last')
                self._salvar_sheet(df, sheet_name='Config')
            elif col_missing:
                self._salvar_sheet(df, sheet_name='Config')

            return df

    def get_config(self, tipo: str = None, chave: str = None) -> pd.DataFrame:
        try:
            df = self._garantir_config_basica()
            if tipo:
                df = df[df['tipo'] == tipo]
            if chave:
                df = df[df['chave'] == chave]
            return df
        except Exception as e:
            print(f"Erro ao ler Config: {e}")
            return pd.DataFrame(columns=COLUNAS_CONFIG)
    
    def get_medicos(self) -> List[str]:
        df = self.get_config(tipo='medico')
        return sorted(df['valor'].tolist()) if not df.empty else []
    
    def get_medico_chaves(self) -> Dict[str, str]:
        df = self.get_config(tipo='medico')
        if df.empty:
            return {}
        return dict(zip(df['chave'], df['valor']))

    def _normalizar_nome_consulta(self, nome: str) -> str:
        return normalize_text(nome).replace(' ', '_')

    def salvar_tipo_consulta_especial(self, nome: str, dia_semana: str, turno: str, medico: str, descricao: str = "") -> None:
        chave = self._normalizar_nome_consulta(nome)
        regra = {
            'nome': nome.strip(),
            'dia_semana': dia_semana.strip(),
            'turno': turno.strip(),
            'medico': medico.strip(),
            'descricao': descricao.strip(),
        }
        df = self._garantir_config_basica()
        idx = df[(df['tipo'] == 'tipo_consulta_especial') & (df['chave'] == chave)].index
        valor = json.dumps(regra, ensure_ascii=False)
        if not idx.empty:
            df.loc[idx[0], 'valor'] = valor
        else:
            nova_linha = pd.DataFrame([{'tipo': 'tipo_consulta_especial', 'chave': chave, 'valor': valor}])
            df = pd.concat([df, nova_linha], ignore_index=True)
        df = df.drop_duplicates(subset=['tipo', 'chave'], keep='last')
        self._salvar_sheet(df, sheet_name='Config')

    def get_tipos_consultas_especiais(self) -> List[Dict[str, str]]:
        df = self.get_config(tipo='tipo_consulta_especial')
        tipos = []
        for _, row in df.iterrows():
            try:
                regra = json.loads(str(row.get('valor', '{}')))
            except Exception:
                regra = {'nome': str(row.get('valor', ''))}
            if isinstance(regra, dict):
                tipos.append({
                    'nome': regra.get('nome', str(row.get('chave', ''))),
                    'dia_semana': regra.get('dia_semana', ''),
                    'turno': regra.get('turno', 'manha'),
                    'medico': regra.get('medico', ''),
                    'descricao': regra.get('descricao', ''),
                })
        return tipos

    def get_tipos_consulta_disponiveis(self) -> List[str]:
        tipos = ['Normal']
        for regra in self.get_tipos_consultas_especiais():
            if regra.get('nome'):
                tipos.append(regra['nome'])
        return tipos

    def get_regra_tipo_consulta(self, tipo: str) -> Dict[str, str]:
        if not tipo or tipo == 'Normal':
            return {}
        regras_default = get_modalidade_rules()
        if tipo in regras_default:
            return regras_default[tipo]
        for regra in self.get_tipos_consultas_especiais():
            if regra.get('nome') == tipo:
                dia_semana = regra.get('dia_semana', '')
                weekday = ['seg', 'ter', 'qua', 'qui', 'sex', 'sab', 'dom'].index(dia_semana) if dia_semana in ['seg', 'ter', 'qua', 'qui', 'sex', 'sab', 'dom'] else 0
                return {
                    'weekday': weekday,
                    'turno': regra.get('turno', 'manha'),
                    'medico': regra.get('medico', ''),
                    'descricao': regra.get('descricao', ''),
                    'dia_semana': dia_semana,
                }
        return {}

    def salvar_datas_indisponiveis(self, medico: str, datas_str: str) -> None:
        if not medico:
            return
        medico_chaves = self.get_medico_chaves()
        chave = ''
        for chave_medico, nome_medico in medico_chaves.items():
            if nome_medico == medico:
                chave = chave_medico
                break
        if not chave:
            return
        df = self._garantir_config_basica()
        idx = df[(df['tipo'] == 'indisponibilidade') & (df['chave'] == chave)].index
        valor = datas_str.strip()
        if not idx.empty:
            df.loc[idx[0], 'valor'] = valor
        else:
            nova_linha = pd.DataFrame([{'tipo': 'indisponibilidade', 'chave': chave, 'valor': valor}])
            df = pd.concat([df, nova_linha], ignore_index=True)
        df = df.drop_duplicates(subset=['tipo', 'chave'], keep='last')
        self._salvar_sheet(df, sheet_name='Config')

    def get_datas_indisponiveis(self, medico: str) -> List[str]:
        if not medico:
            return []
        medico_chaves = self.get_medico_chaves()
        chave = ''
        for chave_medico, nome_medico in medico_chaves.items():
            if nome_medico == medico:
                chave = chave_medico
                break
        if not chave:
            return []
        df = self.get_config(tipo='indisponibilidade', chave=chave)
        if df.empty:
            return []
        return [item.strip() for item in str(df.iloc[0]['valor']).split(',') if item.strip()]

    def _normalizar_modalidade_chave(self, modalidade: str) -> str:
        if not modalidade:
            return ""
        return normalize_text(modalidade).replace(' ', '_')

    def _chave_medico_fallback(self, nome_medico: str) -> str:
        if not nome_medico:
            return ""
        chave = normalize_text(nome_medico).replace(' ', '_')
        return chave if chave.startswith('dr_') else f"dr_{chave}"

    def _gerar_horarios_padrao(self, turno: str) -> List[str]:
        if turno == 'manha':
            inicio, fim = 8, 12
        else:
            inicio, fim = 13, 17
        horarios = []
        for h in range(inicio, fim):
            for m in ['00', '30']:
                horarios.append(f"{h:02d}:{m}")
        return horarios

    def get_modalidade_medico(self, modalidade: str) -> str:
        chave = self._normalizar_modalidade_chave(modalidade)
        if not chave:
            return ""
        df = self.get_config(tipo='modalidade_medico', chave=chave)
        if df.empty:
            return ""
        return df.iloc[0]['valor']

    def get_modalidade_medico_chave(self, modalidade: str) -> str:
        nome_medico = self.get_modalidade_medico(modalidade)
        if not nome_medico:
            return ""
        medico_chaves = self.get_medico_chaves()
        chaves_invertidas = {v: k for k, v in medico_chaves.items()}
        return chaves_invertidas.get(nome_medico, self._chave_medico_fallback(nome_medico))
    
    def get_horarios(self, medico_chave: str, dia_semana: str, turno: str) -> List[str]:
        chave = f"{medico_chave}_{dia_semana}_{turno}"
        df = self.get_config(tipo='horario', chave=chave)
        if df.empty:
            return self._gerar_horarios_padrao(turno)
        horarios_str = df.iloc[0]['valor']
        return [h.strip() for h in horarios_str.split(',') if h.strip()]

    def get_horarios_modalidade(self, modalidade: str, dia_semana: str, turno: str) -> List[str]:
        chave = f"{modalidade}_{dia_semana}_{turno}"
        df = self.get_config(tipo='horario_modalidade', chave=chave)
        if df.empty:
            return self._gerar_horarios_padrao_modalidade(turno)
        horarios_str = df.iloc[0]['valor']
        return [h.strip() for h in horarios_str.split(',') if h.strip()]

    def _gerar_horarios_padrao_modalidade(self, turno: str) -> List[str]:
        if turno == 'manha':
            inicio, fim = 8, 12
        else:
            inicio, fim = 13, 17
        horarios = []
        for h in range(inicio, fim):
            for m in ['00', '30']:
                horarios.append(f"{h:02d}:{m}")
        return horarios

    def salvar_horario_modalidade(self, modalidade: str, dia_semana: str, turno: str, horarios_str: str):
        chave = f"{modalidade}_{dia_semana}_{turno}"
        df = self._garantir_config_basica()
        idx = df[(df['tipo'] == 'horario_modalidade') & (df['chave'] == chave)].index
        if not idx.empty:
            df.loc[idx[0], 'valor'] = horarios_str
        else:
            nova_linha = pd.DataFrame([{'tipo': 'horario_modalidade', 'chave': chave, 'valor': horarios_str}])
            df = pd.concat([df, nova_linha], ignore_index=True)
        df = df.drop_duplicates(subset=['tipo', 'chave'], keep='last')
        self._salvar_sheet(df, sheet_name='Config')
    
    def get_limite_encaixes(self, medico_chave: str, dia_semana: str, turno: str) -> int:
        chave = f"{medico_chave}_{dia_semana}_{turno}"
        df = self.get_config(tipo='espontanea', chave=chave)
        if df.empty:
            return 4
        try:
            return int(df.iloc[0]['valor'])
        except:
            return 4

    def get_horarios_disponiveis(self, medico: str, data: str, tipo_consulta: str = 'Normal') -> List[str]:
        if not medico or not data:
            return []

        medico_chaves = self.get_medico_chaves()
        chaves_invertidas = {v: k for k, v in medico_chaves.items()}
        medico_chave = chaves_invertidas.get(medico, self._chave_medico_fallback(medico))
        if not medico_chave:
            return []

        try:
            data_obj = datetime.strptime(data, '%Y-%m-%d').date()
        except ValueError:
            return []

        dia_semana = ['seg', 'ter', 'qua', 'qui', 'sex', 'sab', 'dom'][data_obj.weekday()]
        turnos = ['manha', 'tarde']
        if tipo_consulta != 'Normal':
            regra = get_modalidade_rules().get(tipo_consulta, {})
            turno_especifico = regra.get('turno', 'manha')
            turnos = [turno_especifico]

        agenda_df = self.get_agenda(data=data, medico=medico, status='ATIVO')
        horarios_disponiveis: List[str] = []
        for turno in turnos:
            horarios = self.get_horarios(medico_chave, dia_semana, turno)
            if tipo_consulta != 'Normal':
                ocupados = agenda_df[
                    (agenda_df['tipo_consulta'] == tipo_consulta) & (agenda_df['medico'] == medico)
                ]['hora'].tolist() if not agenda_df.empty else []
            else:
                ocupados = agenda_df[
                    (agenda_df['tipo_consulta'] == 'Normal') & (agenda_df['medico'] == medico)
                ]['hora'].tolist() if not agenda_df.empty else []
            horarios_disponiveis.extend([h for h in horarios if h not in ocupados])

        return horarios_disponiveis
    
    def salvar_config(self, df_config: pd.DataFrame):
        if df_config is None:
            df_config = pd.DataFrame(columns=COLUNAS_CONFIG)

        df = df_config.copy()
        for col in COLUNAS_CONFIG:
            if col not in df.columns:
                df[col] = ''
        df = df.dropna(how='all')
        df['tipo'] = df['tipo'].fillna('').astype(str)
        df['chave'] = df['chave'].fillna('').astype(str)
        df['valor'] = df['valor'].fillna('').astype(str)
        df = df.drop_duplicates(subset=['tipo', 'chave'], keep='last')

        self._salvar_sheet(df, sheet_name='Config')
    
    # ==================== AGENDA ====================
    
    def get_proximo_id_agenda(self) -> int:
        try:
            df = self._obter_sheet("Agenda", COLUNAS_AGENDA)
            if df.empty or 'id' not in df.columns:
                return 1
            ids = pd.to_numeric(df['id'], errors='coerce').dropna()
            return int(ids.max() + 1) if not ids.empty else 1
        except Exception:
            return 1
    
    def get_agenda(self, data: str = None, medico: str = None, status: str = 'ATIVO') -> pd.DataFrame:
        try:
            df = self._obter_sheet("Agenda", COLUNAS_AGENDA)
            if df.empty:
                return pd.DataFrame(columns=COLUNAS_AGENDA)
            if data:
                df = df[df['data'] == str(data)]
            if medico:
                df = df[df['medico'] == str(medico)]
            if status:
                df = df[df['status'] == str(status)]
            return df.sort_values('hora', na_position='last')
        except Exception as e:
            print(f"Erro ao ler Agenda: {e}")
            return pd.DataFrame(columns=COLUNAS_AGENDA)
    
    def get_agenda_do_dia(self) -> pd.DataFrame:
        hoje = date_str()
        return self.get_agenda(data=hoje, status='ATIVO')
    
    def agendar_consulta(self, paciente: str, medico: str, tipo_consulta: str,
                         data: str, hora: str, encaixe: bool = False,
                         observacao: str = "", usuario: str = "Sistema") -> int:
        with self._lock:
            df = self._obter_sheet("Agenda", COLUNAS_AGENDA)
            novo_id = self.get_proximo_id_agenda()
            timestamp = timestamp_str()
            
            novo_registro = {
                'id': str(novo_id),
                'paciente': paciente.strip(),
                'medico': medico.strip(),
                'tipo_consulta': tipo_consulta.strip(),
                'data': str(data),
                'hora': str(hora),
                'encaixe': 'TRUE' if encaixe else 'FALSE',
                'status': 'ATIVO',
                'compareceu': 'FALSE',
                'quando_confirmou_presenca': '',
                'observacao': observacao.strip(),
                'criado_em': timestamp,
                'criado_por': usuario.strip()
            }
            
            for col in COLUNAS_AGENDA:
                if col not in df.columns:
                    df[col] = ''
            
            novo_df = pd.DataFrame([novo_registro])
            df = pd.concat([df, novo_df], ignore_index=True)
            
            self._salvar_sheet(df, sheet_name='Agenda')
            
            tipo_acao = ACAO_ENCAIXAR if encaixe else ACAO_AGENDAR
            detalhes = f"{tipo_consulta} - {paciente} - {medico} - {data} {hora}"
            self.registrar_log(tipo_acao, detalhes, usuario)
            
            return novo_id
    
    def cancelar_consulta(self, id_agenda: int, usuario: str = "Sistema"):
        with self._lock:
            df = self._obter_sheet("Agenda", COLUNAS_AGENDA)
            id_str = str(id_agenda)
            idx = df[df['id'] == id_str].index
            if idx.empty:
                return
            df.loc[idx, 'status'] = 'CANCELADO'
            self._salvar_sheet(df, sheet_name='Agenda')
            paciente = df.loc[idx[0], 'paciente']
            self.registrar_log(ACAO_CANCELAR, f"ID {id_agenda} - {paciente}", usuario)

    def editar_consulta(self, id_agenda: int, campos: dict, usuario: str = "Sistema"):
        with self._lock:
            df = self._obter_sheet("Agenda", COLUNAS_AGENDA)
            id_str = str(id_agenda)
            idx = df[df['id'] == id_str].index

            if idx.empty:
                raise ValueError(f"Consulta ID {id_agenda} não encontrada.")

            campos_permitidos = {'paciente', 'medico', 'tipo_consulta', 'data', 'hora', 'encaixe', 'observacao'}
            alteracoes = []

            for campo, valor in campos.items():
                if campo not in campos_permitidos:
                    continue
                valor_str = str(valor).strip()
                valor_anterior = str(df.loc[idx[0], campo]).strip() if campo in df.columns else ''
                if valor_str != valor_anterior:
                    df.loc[idx, campo] = valor_str
                    alteracoes.append(f"{campo}: '{valor_anterior}' → '{valor_str}'")

            if not alteracoes:
                return

            self._salvar_sheet(df, sheet_name='Agenda')

            paciente = df.loc[idx[0], 'paciente']
            detalhes = f"ID {id_agenda} - {paciente} | " + " | ".join(alteracoes)
            self.registrar_log(ACAO_EDITAR, detalhes, usuario)

    def marcar_presenca(self, id_agenda: int, presente: bool = True, usuario: str = "Sistema"):
        with self._lock:
            df = self._obter_sheet("Agenda", COLUNAS_AGENDA)
            id_str = str(id_agenda)
            idx = df[df['id'] == id_str].index
            if idx.empty:
                return
            df.loc[idx, 'compareceu'] = 'TRUE' if presente else 'FALSE'
            df.loc[idx, 'quando_confirmou_presenca'] = timestamp_str() if presente else ''
            self._salvar_sheet(df, sheet_name='Agenda')
            acao = ACAO_MARCAR_PRESENCA if presente else ACAO_DESMARCAR_PRESENCA
            paciente = df.loc[idx[0], 'paciente']
            self.registrar_log(acao, f"ID {id_agenda} - {paciente}", usuario)
    
    def contar_encaixes_usados(self, medico: str, data: str) -> int:
        df = self.get_agenda(data=data, medico=medico, status='ATIVO')
        if df.empty:
            return 0
        return len(df[df['encaixe'] == 'TRUE'])
    
    # ==================== RECEITAS ====================
    
    def get_proximo_id_receita(self) -> int:
        try:
            df = self._obter_sheet("Receitas", COLUNAS_RECEITAS)
            if df.empty or 'id' not in df.columns:
                return 1
            ids = pd.to_numeric(df['id'], errors='coerce').dropna()
            return int(ids.max() + 1) if not ids.empty else 1
        except Exception:
            return 1
    
    def get_receitas_pendentes(self) -> pd.DataFrame:
        try:
            df = self._obter_sheet("Receitas", COLUNAS_RECEITAS)
            if df.empty:
                return pd.DataFrame(columns=COLUNAS_RECEITAS)
            return df[df['status'] == 'PENDENTE'].sort_values('data_pedido', ascending=False)
        except Exception:
            return pd.DataFrame(columns=COLUNAS_RECEITAS)
    
    def get_todas_receitas(self) -> pd.DataFrame:
        try:
            df = self._obter_sheet("Receitas", COLUNAS_RECEITAS)
            if df.empty:
                return pd.DataFrame(columns=COLUNAS_RECEITAS)
            return df.sort_values('data_pedido', ascending=False)
        except Exception:
            return pd.DataFrame(columns=COLUNAS_RECEITAS)
    
    def pedir_receita(self, paciente: str, observacao: str = "", usuario: str = "Sistema") -> int:
        with self._lock:
            df = self._obter_sheet("Receitas", COLUNAS_RECEITAS)
            novo_id = self.get_proximo_id_receita()
            timestamp = timestamp_str()
            
            novo_registro = {
                'id': str(novo_id),
                'paciente': paciente.strip(),
                'status': 'PENDENTE',
                'quem_retirou': '',
                'observacao': observacao.strip(),
                'data_pedido': timestamp,
                'data_retirada': ''
            }
            
            for col in COLUNAS_RECEITAS:
                if col not in df.columns:
                    df[col] = ''
            
            novo_df = pd.DataFrame([novo_registro])
            df = pd.concat([df, novo_df], ignore_index=True)
            
            self._salvar_sheet(df, sheet_name='Receitas')
            self.registrar_log(ACAO_PEDIR_RECEITA, f"{paciente}", usuario)
            return novo_id
    
    def retirar_receita(self, id_receita: int, quem_retirou: str, 
                        observacao: str = "", usuario: str = "Sistema"):
        with self._lock:
            df = self._obter_sheet("Receitas", COLUNAS_RECEITAS)
            timestamp = timestamp_str()
            
            id_receita_str = str(id_receita)
            idx = df[df['id'] == id_receita_str].index
            if idx.empty:
                return

            df.loc[idx, 'status'] = 'RETIRADA'
            df.loc[idx, 'quem_retirou'] = quem_retirou.strip()
            df.loc[idx, 'data_retirada'] = timestamp
            if observacao:
                obs_atual = str(df.loc[idx[0], 'observacao'] or '')
                df.loc[idx, 'observacao'] = f"{obs_atual} | Retirada: {observacao.strip()}".strip(" |")
            
            self._salvar_sheet(df, sheet_name='Receitas')
            paciente = df.loc[idx[0], 'paciente']
            self.registrar_log(ACAO_RETIRAR_RECEITA, f"ID {id_receita} - {paciente} - Retirado por: {quem_retirou}", usuario)
    
    # ==================== DATA DE RETIRADA DE RECEITAS ====================
    
    def calcular_data_retirada_receita(self, data_pedido_str: str = None) -> str:
        """
        Calcula a data prevista de retirada da receita com base na regra:
        - Pedido na SEGUNDA → retirada na SEXTA da MESMA semana
        - Pedido de TERÇA a SEXTA → retirada na SEXTA da PRÓXIMA semana
        """
        if data_pedido_str:
            try:
                data_pedido = datetime.strptime(data_pedido_str[:10], "%Y-%m-%d")
            except Exception:
                data_pedido = now()
        else:
            data_pedido = now()
        
        dia_semana = data_pedido.weekday()  # 0=segunda, 1=terça, ..., 4=sexta
        
        if dia_semana == 0:  # Segunda-feira
            dias_para_sexta = 4
            data_retirada = data_pedido + timedelta(days=dias_para_sexta)
        elif dia_semana <= 4:  # Terça a Sexta
            dias_para_sexta = (4 - dia_semana) + 7
            data_retirada = data_pedido + timedelta(days=dias_para_sexta)
        else:  # Sábado ou Domingo
            dias_para_sexta = (4 - dia_semana) + 7
            data_retirada = data_pedido + timedelta(days=dias_para_sexta)
        
        return data_retirada.strftime("%Y-%m-%d")
    
    def get_data_retirada_formatada(self, data_pedido_str: str) -> str:
        data_retirada = self.calcular_data_retirada_receita(data_pedido_str)
        try:
            data_obj = datetime.strptime(data_retirada, "%Y-%m-%d")
            return data_obj.strftime("%d/%m/%Y")
        except Exception:
            return data_retirada
    
    def get_receitas_pendentes_com_retirada(self) -> List[Dict]:
        pendentes = self.get_receitas_pendentes()
        if pendentes.empty:
            return []
        
        resultado = []
        for _, receita in pendentes.iterrows():
            item = receita.to_dict()
            data_pedido = str(receita.get('data_pedido', ''))
            item['data_retirada_prevista'] = self.calcular_data_retirada_receita(data_pedido)
            item['data_retirada_formatada'] = self.get_data_retirada_formatada(data_pedido)
            resultado.append(item)
        
        return resultado
    
    # ==================== LOG ====================
    
    def get_proximo_id_log(self) -> int:
        try:
            df = self._obter_sheet("Log", COLUNAS_LOG)
            if df.empty or 'id' not in df.columns:
                return 1
            ids = pd.to_numeric(df['id'], errors='coerce').dropna()
            return int(ids.max() + 1) if not ids.empty else 1
        except Exception:
            return 1
    
    def registrar_log(self, acao: str, detalhes: str = "", usuario: str = "Sistema"):
        with self._lock:
            try:
                df = self._obter_sheet("Log", COLUNAS_LOG)
                novo_id = self.get_proximo_id_log()
                timestamp = timestamp_str()
                
                novo_registro = {
                    'id': str(novo_id),
                    'timestamp': timestamp,
                    'acao': acao,
                    'detalhes': str(detalhes),
                    'usuario': str(usuario)
                }
                
                for col in COLUNAS_LOG:
                    if col not in df.columns:
                        df[col] = ''
                
                novo_df = pd.DataFrame([novo_registro])
                df = pd.concat([df, novo_df], ignore_index=True)
                
                self._salvar_sheet(df, sheet_name='Log')
            except Exception as e:
                print(f"Erro ao registrar log: {e}")
    
    # ==================== BUSCA GLOBAL ====================
    
    def buscar_global(self, termo: str) -> Dict[str, List[Dict]]:
        resultados = {
            'agenda': [],
            'receitas': [],
        }
        
        if not termo or not termo.strip():
            return resultados
        
        termo_limpo = termo.strip()
        
        try:
            df_agenda = self._obter_sheet("Agenda", COLUNAS_AGENDA)
            if not df_agenda.empty:
                for _, row in df_agenda.iterrows():
                    campos_busca = [
                        row.get('paciente', ''),
                        row.get('medico', ''),
                        row.get('tipo_consulta', ''),
                        row.get('observacao', ''),
                        row.get('data', ''),
                    ]
                    if any(search_match(termo_limpo, str(c)) for c in campos_busca):
                        resultados['agenda'].append({
                            'id': row.get('id'),
                            'paciente': row.get('paciente', ''),
                            'medico': row.get('medico', ''),
                            'tipo': row.get('tipo_consulta', ''),
                            'data': row.get('data', ''),
                            'hora': row.get('hora', ''),
                            'status': row.get('status', ''),
                            'encaixe': row.get('encaixe', 'FALSE'),
                            'compareceu': row.get('compareceu', 'FALSE'),
                            'observacao': row.get('observacao', ''),
                        })
        except Exception as e:
            print(f"Erro na busca em Agenda: {e}")
        
        try:
            df_receitas = self._obter_sheet("Receitas", COLUNAS_RECEITAS)
            if not df_receitas.empty:
                for _, row in df_receitas.iterrows():
                    campos_busca = [
                        row.get('paciente', ''),
                        row.get('quem_retirou', ''),
                        row.get('observacao', ''),
                    ]
                    if any(search_match(termo_limpo, str(c)) for c in campos_busca):
                        resultados['receitas'].append({
                            'id': row.get('id'),
                            'paciente': row.get('paciente', ''),
                            'status': row.get('status', ''),
                            'quem_retirou': row.get('quem_retirou', ''),
                            'data_pedido': row.get('data_pedido', ''),
                            'observacao': row.get('observacao', ''),
                        })
        except Exception as e:
            print(f"Erro na busca em Receitas: {e}")
        
        return resultados
